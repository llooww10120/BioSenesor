import serial
from openpyxl import Workbook
from openpyxl import load_workbook

import matplotlib.pyplot as plt
import time
ser = serial.Serial("COM3",9600)
name="text1.xlsx"
wb=Workbook()
ws=wb.active
ws["A1"]="Ohm"
while True:
    try:
        content=float(ser.readline().decode("utf-8")[:-2])
        ws.append([time.strftime("%Y-%m-%d %H:%M:%S",time.localtime()),content])
        print(content)
    except:
        break

wb.save(name)
wb=load_workbook(name)
sheet=wb["Sheet"]
a=sheet["A"]
b=sheet["B"]
ax=[]
by=[]
for cell in a:
    ax.append(cell.value[-8:])

for cell in b:
    by.append(cell.value)


fig = plt.figure()
axx=fig.add_subplot(111)
plt.plot(ax,by,color="blue",linewidth=2,marker="o")
plt.savefig(name+".png")
plt.show()
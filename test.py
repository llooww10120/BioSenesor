import serial
from openpyxl import Workbook
from openpyxl import load_workbook
import time

ser = serial.Serial("/dev/ttyACM0",9600)
name="text1.xlsx"
wb=Workbook()
ws=wb.active
localtime=time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())
rownum=1
while localtime <= "2021-08-30 10:45:00":
    localtime=time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())
    # content=str(ser.readline().decode("utf-8")[:-2])
    # print(str(ser.readline().decode().replace('\n','')=="ALL DONE\r"))
    if(str(ser.readline().decode().replace('\n',''))=="ALL DONE\r"):
        ws.cell(row=rownum,column=1,value=localtime)
        for i in range(2,258):
            ws.cell(row=rownum,column=i,value=str(ser.readline().decode().replace('\n','')))
        rownum+=1
    # print(str(ser.readline().decode("utf-8")[:-2]))
print("end")
wb.save(name)

import serial
from openpyxl import Workbook
from openpyxl import load_workbook
import time

ser = serial.Serial("COM3",115200)
localtime=time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())
name=localtime[:10]+".csv"
wb=Workbook()
ws=wb.active
rownum=1

while localtime <= "2021-09-01 17:50:00":
    localtime=time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())
    # content=str(ser.readline().decode("utf-8")[:-2])
    # print(str(ser.readline().decode().replace('\n','')=="ALL DONE\r"))
    if(str(ser.readline().decode().replace('\n',''))=="ALL DONE\r"):
        ws.cell(row=rownum,column=1,value=localtime)
        for i in range(2,258):
            ws.cell(row=rownum,column=i,value=str(ser.readline().decode().replace('\n','')))
        rownum+=1

    # print(str(ser.readline().decode("utf-8")[:-2]))
print("end")
wb.save(name)
